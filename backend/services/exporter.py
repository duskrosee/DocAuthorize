import os
import shutil
import smtplib
from email.message import EmailMessage
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from PySide6.QtCore import QSettings
from database.db import get_documents_by_nips


def save_contracts_only(parsed_data_list: list, root_dir: Path) -> bool:
    """Only saving and organizing files into folders (WITHOUT sending mail)."""
    if not parsed_data_list:
        return False

    for item in parsed_data_list:
        target_folder = root_dir / "documents" / ("signed" if item.get("status") == "Signed" else "pending")
        target_folder.mkdir(parents=True, exist_ok=True)
        
        target_path = target_folder / item["file_name"]
        try:
            if Path(item["file_path"]).exists():
                shutil.copy2(item["file_path"], target_path)
        except Exception as e:
            print(f"[Exporter] Failed to copy {item['file_path']}: {e}")

    print("[Exporter]: Files successfully saved.")
    return True


def send_new_documents_report(parsed_data_list: list, root_dir: Path) -> bool:
    """FOR THE DOCUMENTS TAB: Sends a report ONLY on transferred (new) files."""
    if not parsed_data_list:
        return False

    reports_dir = root_dir / "reports"
    reports_dir.mkdir(exist_ok=True)
    excel_path = reports_dir / "New_Contracts_Report.xlsx"

    return _generate_and_send_excel(parsed_data_list, excel_path)


def send_client_history_report(nip: str, root_dir: Path) -> bool:
    """FOR THE CLIENTS TAB: Sends the complete document history for a specific firm via NIP."""
    if not nip or nip == "---":
        return False

    all_docs = get_documents_by_nips([nip])
    if not all_docs:
        print(f"[Exporter] No history found for NIP: {nip}")
        return False

    reports_dir = root_dir / "reports"
    reports_dir.mkdir(exist_ok=True)
    excel_path = reports_dir / f"Client_History_{nip}.xlsx"

    return _generate_and_send_excel(all_docs, excel_path)


def _generate_and_send_excel(docs_list: list, excel_path: Path) -> bool:
    """Internal function: creates Excel and calls email sending."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contracts"

    headers = ["Filename", "Company", "NIP", "Price", "Contract Date", "Manager"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    pdf_files_to_send = []

    for item in docs_list:
        ws.append([
            item.get("file_name", "---"),
            item.get("company_name", "---"),
            item.get("nip", "---"),
            item.get("price", "---"),
            item.get("contract_date", "---"),
            item.get("manager", "---")
        ])

        fpath = Path(item.get("file_path", ""))
        if fpath.exists():
            pdf_files_to_send.append(fpath)

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = max(max_len + 4, 12)

    wb.save(excel_path)
    return send_email_to_boss(excel_path, pdf_files_to_send)


def send_email_to_boss(excel_path: Path, pdf_files: list) -> bool:
    """Sends an email with attached Excel and PDF files through Outlook (Office 365)."""
    settings = QSettings("DocAuthorize", "ContractManager")
    
    sender_email = settings.value("corp_email", "biuro@pmdigital.pl")
    sender_password = settings.value("corp_password", "")
    boss_email = settings.value("boss_email", "")

    if not sender_password or not boss_email or not sender_email:
        print("[Email Error] Credentials missing! Please set Email, Password and Boss Email in Settings.")
        return False

    smtp_server = "smtp.office365.com"
    smtp_port = 587

    try:
        msg = EmailMessage()
        msg["Subject"] = "Raport: Umowy Klientów (DocAuthorize)"
        msg["From"] = sender_email
        msg["To"] = boss_email
        msg.set_content("Dzień dobry,\n\nW załączniku przesyłam zbiorczy raport umów oraz pliki PDF.\n\nPozdrawiam.")

        with open(excel_path, "rb") as f:
            msg.add_attachment(
                f.read(),
                maintype="application",
                subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=excel_path.name
            )

        for pdf_path in pdf_files:
            if pdf_path.exists():
                try:
                    with open(pdf_path, "rb") as f:
                        msg.add_attachment(
                            f.read(), maintype="application", subtype="pdf", filename=pdf_path.name
                        )
                except Exception as e:
                    print(f"[Email Attachment Error] {pdf_path}: {e}")

        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.send_message(msg)

        print(f"[Email Success]: Report successfully sent to {boss_email} from {sender_email}")
        return True
    except Exception as e:
        print(f"[Email Error]: {e}")
        return False


    #.